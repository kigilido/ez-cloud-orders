"""
Management command: import_spruce
Usage: python manage.py import_spruce <path_to_excel_file>

Imports customer eligibility data from the Spruce Excel file into the database.
Columns expected: Family #, Medicaid ID, Last Name, First Name,
                  Address, City State Zip, Phone Number, Valid Until, Notes
"""

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.customers.models import Customer


def normalize_phone(raw: str) -> str:
    """Strip all non-digit characters, return digits only."""
    if not raw:
        return ""
    return re.sub(r"\D", "", str(raw))


class Command(BaseCommand):
    help = "Import customers from a Spruce Excel (.xlsx) file"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Spruce .xlsx file")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview what would be imported without saving to the database",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is required. Run: pip install openpyxl"
            )

        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        dry_run = options["dry_run"]
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no data will be saved."))

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Read header row to map column names → indices
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.stdout.write(f"Detected columns: {headers}")

        col = {name: idx for idx, name in enumerate(headers)}

        required = ["Family #", "Medicaid ID", "Last Name", "First Name",
                    "Address", "City, State Zip", "Phone Number", "Valid Until"]
        missing = [r for r in required if r not in col]
        if missing:
            raise CommandError(f"Missing required columns: {missing}\nFound: {headers}")

        created = updated = skipped = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue

            family_number = str(row[col["Family #"]]).strip() if row[col["Family #"]] is not None else ""
            medicaid_id   = str(row[col["Medicaid ID"]]).strip() if row[col["Medicaid ID"]] is not None else ""
            last_name     = str(row[col["Last Name"]]).strip() if row[col["Last Name"]] is not None else ""
            first_name    = str(row[col["First Name"]]).strip() if row[col["First Name"]] is not None else ""
            address       = str(row[col["Address"]]).strip() if row[col["Address"]] is not None else ""
            city_state_zip = str(row[col["City, State Zip"]]).strip() if row[col["City, State Zip"]] is not None else ""
            phone_raw     = row[col["Phone Number"]]
            valid_until   = row[col["Valid Until"]]
            notes_col     = col.get("Notes")
            notes         = str(row[notes_col]).strip() if notes_col is not None and row[notes_col] is not None else ""

            # Validate required fields
            if not family_number or not medicaid_id:
                self.stdout.write(self.style.WARNING(f"Row {row_num}: missing Family # or Medicaid ID — skipped"))
                skipped += 1
                continue

            if valid_until is None:
                self.stdout.write(self.style.WARNING(f"Row {row_num}: missing Valid Until date — skipped"))
                skipped += 1
                continue

            # valid_until comes as a datetime from openpyxl; convert to date
            if hasattr(valid_until, "date"):
                valid_until = valid_until.date()

            phone = normalize_phone(str(phone_raw)) if phone_raw else ""

            if dry_run:
                self.stdout.write(
                    f"  Row {row_num}: {first_name} {last_name} | "
                    f"Phone: {phone} | Valid until: {valid_until}"
                )
                created += 1
                continue

            obj, was_created = Customer.objects.update_or_create(
                medicaid_id=medicaid_id,
                defaults={
                    "family_number": family_number,
                    "first_name": first_name,
                    "last_name": last_name,
                    "address": address,
                    "city_state_zip": city_state_zip,
                    "phone_number": phone,
                    "valid_until": valid_until,
                    "notes": notes,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        if dry_run:
            self.stdout.write(self.style.SUCCESS(f"\nDry run complete — would import {created} customers."))
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"\nImport complete — created: {created}, updated: {updated}, skipped: {skipped}"
                )
            )
